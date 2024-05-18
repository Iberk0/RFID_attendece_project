import serial
import openpyxl
from openpyxl import Workbook
import os

# Configure the serial port (adjust 'COM3' and baud rate as needed)
ser = serial.Serial('COM8', 9600, timeout=1)

def load_or_create_workbook(filename):
    # Check if the file exists and is a valid Excel file
    if os.path.exists(filename):
        try:
            workbook = openpyxl.load_workbook(filename)
            sheet = workbook.active
            return workbook, sheet
        except openpyxl.utils.exceptions.InvalidFileException:
            print(f"{filename} is not a valid Excel file. Creating a new file.")
    
    # Create a new workbook and sheet
    workbook = Workbook()
    sheet = workbook.active
    sheet['A1'] = 'UID'
    sheet['B1'] = 'Name'
    sheet['C1'] = 'Surname'
    workbook.save(filename)
    return workbook, sheet

def is_valid_uid(uid):
    # Add your validation logic here
    # For example, check if the UID length is 10 characters
    return len(uid) == 10

workbook, sheet = load_or_create_workbook('liste.xlsx')

def write_to_excel(uid, name, surname):
    # Find the next empty row
    next_row = sheet.max_row + 1
    sheet[f'A{next_row}'] = uid
    sheet[f'B{next_row}'] = name
    sheet[f'C{next_row}'] = surname
    workbook.save('liste.xlsx')

print("Listening for RFID data...")

try:
    while True:
        if ser.in_waiting > 0:
            uid = ser.readline().decode('utf-8').strip()
            if is_valid_uid(uid):
                print(f"Valid UID received: {uid}")
                name = input("Enter Name: ")
                surname = input("Enter Surname: ")
                write_to_excel(uid, name, surname)
                print("Data saved to Excel.\n")
            else:
                print(f"Invalid UID received: {uid}. Waiting for a valid UID...")

except KeyboardInterrupt:
    print("Exiting...")

finally:
    ser.close()
    print("Serial port closed.")
